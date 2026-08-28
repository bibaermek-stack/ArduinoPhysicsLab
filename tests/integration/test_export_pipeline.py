"""Экспорт ағыны: бір сессия → CSV / Excel / PDF + IExporter factory."""

import csv
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from core.exceptions import ExportError
from domain.entities.experiment_session import ExperimentSession
from domain.entities.measurement import Measurement
from domain.interfaces.i_exporter import IExporter
from domain.services.csv_exporter import CSVExporter
from domain.services.excel_exporter import ExcelExporter
from domain.services.pdf_exporter import PDFExporter
from infrastructure.export.exporter_factory import create_exporter


def _session_with_measurements() -> ExperimentSession:
    started = datetime(2026, 3, 1, 12, 0, tzinfo=timezone.utc)
    session = ExperimentSession(id="s-export", experiment_id="E02", started_at=started)
    session.add_measurement(
        Measurement(
            timestamp=started,
            values={"voltage": 5.024, "current": 0.218},
            experiment_id="E02",
            derived_values={"power": 1.095},
        )
    )
    session.add_measurement(
        Measurement(
            timestamp=started.replace(second=1),
            values={"voltage": 5.100, "current": 0.200},
            experiment_id="E02",
            derived_values={"power": 1.020},
        )
    )
    return session


def test_exporters_implement_iexporter() -> None:
    assert isinstance(CSVExporter(), IExporter)
    assert isinstance(ExcelExporter(), IExporter)
    assert isinstance(PDFExporter(), IExporter)


def test_empty_session_is_rejected_by_all_formats(tmp_path: Path) -> None:
    session = ExperimentSession(
        id="empty",
        experiment_id="E02",
        started_at=datetime.now(timezone.utc),
    )

    assert CSVExporter().export(session, str(tmp_path / "a.csv")) is False
    assert ExcelExporter().export(session, str(tmp_path / "a.xlsx")) is False
    assert PDFExporter().export(session, str(tmp_path / "a.pdf")) is False
    assert not (tmp_path / "a.csv").exists()
    assert not (tmp_path / "a.xlsx").exists()
    assert not (tmp_path / "a.pdf").exists()


def test_same_session_exports_to_csv_excel_and_pdf(tmp_path: Path) -> None:
    session = _session_with_measurements()
    csv_path = tmp_path / "export.csv"
    xlsx_path = tmp_path / "export.xlsx"
    pdf_path = tmp_path / "export.pdf"

    assert create_exporter("csv").export(session, csv_path) is True
    assert create_exporter("xlsx").export(session, xlsx_path) is True
    assert create_exporter("pdf").export(session, pdf_path) is True

    assert csv_path.exists()
    assert xlsx_path.exists()
    assert pdf_path.exists()

    with open(csv_path, encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[0] == ["No", "Time(s)", "Voltage(V)", "Current(A)", "Power(W)"]
    assert rows[1][2] == "5.024"
    assert rows[2][3] == "0.200"

    workbook = load_workbook(xlsx_path)
    assert "Measurements" in workbook.sheetnames
    assert workbook["Measurements"]["C2"].value == pytest.approx(5.024)


def test_factory_aliases_and_unknown_format() -> None:
    assert type(create_exporter("CSV")).__name__ == "CSVExporter"
    assert type(create_exporter(".pdf")).__name__ == "PDFExporter"
    assert type(create_exporter("excel")).__name__ == "ExcelExporter"

    with pytest.raises(ExportError, match="Белгісіз экспорт форматы"):
        create_exporter("json")
