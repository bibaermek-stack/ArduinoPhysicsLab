"""ExcelExporter үшін юнит-тесттер."""

from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from domain.entities.experiment_session import ExperimentSession
from domain.entities.measurement import Measurement
from domain.services.excel_exporter import ExcelExporter


def _make_measurement(
    values: dict[str, float] | None = None,
    derived_values: dict[str, float] | None = None,
    timestamp: datetime | None = None,
) -> Measurement:
    return Measurement(
        timestamp=timestamp or datetime.now(timezone.utc),
        values=values or {},
        experiment_id="E02",
        derived_values=derived_values or {},
    )


def _make_session(started_at: datetime | None = None) -> ExperimentSession:
    return ExperimentSession(
        id="s1", experiment_id="E02", started_at=started_at or datetime.now(timezone.utc)
    )


def test_empty_session_returns_false_and_creates_no_file(tmp_path: Path) -> None:
    session = _make_session()
    output_path = tmp_path / "export.xlsx"

    result = ExcelExporter().export(session, str(output_path))

    assert result is False
    assert not output_path.exists()


def test_file_is_created(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    result = ExcelExporter().export(session, str(output_path))

    assert result is True
    assert output_path.exists()


def test_sheet_name_is_measurements(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    workbook = load_workbook(output_path)
    assert "Measurements" in workbook.sheetnames


def test_header_row_is_correct(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    header = [cell.value for cell in sheet[1]]
    assert header == ["№", "Уақыт (сек)", "Кернеу (V)", "Ток (A)", "Қуат (W)"]


def test_one_measurement_is_exported(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(
        _make_measurement(
            values={"voltage": 5.024, "current": 0.218}, derived_values={"power": 1.095}
        )
    )
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    row = [cell.value for cell in sheet[2]]
    assert row[0] == 1
    assert row[1] == 0.0
    assert row[2] == 5.024
    assert row[3] == 0.218
    assert row[4] == 1.095


def test_many_measurements_are_exported_in_order(tmp_path: Path) -> None:
    started_at = datetime.now(timezone.utc)
    session = _make_session(started_at)
    for i in range(1, 4):
        session.add_measurement(
            _make_measurement(
                values={"voltage": float(i)},
                timestamp=started_at + timedelta(seconds=i),
            )
        )
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    assert sheet.max_row == 4  # header + 3
    voltages = [sheet.cell(row=r, column=3).value for r in range(2, 5)]
    assert voltages == [1.0, 2.0, 3.0]


def test_row_numbering_starts_at_one(tmp_path: Path) -> None:
    session = _make_session()
    for i in range(3):
        session.add_measurement(_make_measurement(values={"voltage": float(i)}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    numbers = [sheet.cell(row=r, column=1).value for r in range(2, 5)]
    assert numbers == [1, 2, 3]


def test_missing_values_are_empty_cells(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    assert sheet.cell(row=2, column=4).value is None  # current
    assert sheet.cell(row=2, column=5).value is None  # power


def test_time_number_format(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    assert sheet.cell(row=2, column=2).number_format == "0.00"


def test_value_cells_are_numeric_with_three_decimal_format(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(
        _make_measurement(
            values={"voltage": 5.0, "current": 0.2}, derived_values={"power": 1.0}
        )
    )
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    for column in (3, 4, 5):
        cell = sheet.cell(row=2, column=column)
        assert isinstance(cell.value, (int, float))
        assert not isinstance(cell.value, str)
        assert cell.number_format == "0.000"


def test_freeze_panes_is_a2(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    assert sheet.freeze_panes == "A2"


def test_auto_filter_is_set(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    sheet = load_workbook(output_path)["Measurements"]
    assert sheet.auto_filter.ref is not None


def test_experiment_info_sheet_exists(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path = tmp_path / "export.xlsx"

    ExcelExporter().export(session, str(output_path))

    workbook = load_workbook(output_path)
    assert "Experiment Info" in workbook.sheetnames
    info_sheet = workbook["Experiment Info"]
    labels = [row[0].value for row in info_sheet.iter_rows(min_col=1, max_col=1)]
    assert "Experiment ID" in labels
    assert "Measurement count" in labels


def test_invalid_path_returns_false(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    invalid_path = tmp_path / "no_such_directory" / "export.xlsx"

    result = ExcelExporter().export(session, str(invalid_path))

    assert result is False


def test_xlsx_extension_is_added_automatically(tmp_path: Path) -> None:
    session = _make_session()
    session.add_measurement(_make_measurement(values={"voltage": 5.0}))
    output_path_without_extension = tmp_path / "export"

    result = ExcelExporter().export(session, str(output_path_without_extension))

    assert result is True
    assert (tmp_path / "export.xlsx").exists()
