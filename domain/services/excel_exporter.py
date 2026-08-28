"""excel_exporter — ExperimentSession ішіндегі Measurement тарихын Excel
(.xlsx) жұмыс кітабына экспорттайтын domain сервисі.

Экспорттың жалғыз дереккөзі — ``ExperimentSession.measurements``. Бұл
сервис ``MeasurementTableWidget``-пен немесе ``LiveGraphWidget``-пен
ешбір байланысы жоқ (UI бұл екеуінен деректі оқымайды).
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domain.entities.experiment_session import ExperimentSession
from domain.interfaces.i_exporter import IExporter

_MEASUREMENTS_SHEET_NAME = "Measurements"
_INFO_SHEET_NAME = "Experiment Info"

_HEADER = ("№", "Уақыт (сек)", "Кернеу (V)", "Ток (A)", "Қуат (W)")
_TIME_NUMBER_FORMAT = "0.00"
_VALUE_NUMBER_FORMAT = "0.000"

# (Measurement ішіндегі кілт, баған нөмірі) — Voltage/Current/Power ретімен.
_VALUE_COLUMNS: tuple[tuple[str, int], ...] = (
    ("voltage", 3),
    ("current", 4),
    ("power", 5),
)

_COLUMN_WIDTHS: dict[int, float] = {1: 6, 2: 14, 3: 14, 4: 14, 5: 14}


class ExcelExporter(IExporter):
    """``ExperimentSession``-ды жеңіл форматталған .xlsx жұмыс кітабына
    жазатын сервис.
    """

    def export(self, session: ExperimentSession, output_path: str | Path) -> bool:
        """``session.measurements``-ті ``output_path`` файлына .xlsx
        жұмыс кітабы ретінде жазады.

        Session бос болса, файл жасалмайды және ``False`` қайтарылады.
        Кеңейтім ``.xlsx`` болмаса, автоматты түрде қосылады/ауыстырылады.
        Кез келген қате (parent directory жоқ, рұқсат жоқ, т.б.) ұсталады,
        ешбір exception сыртқа шықпайды.
        """
        if not session.measurements:
            return False

        try:
            path = Path(output_path)
            if path.suffix.lower() != ".xlsx":
                path = path.with_suffix(".xlsx")

            workbook = Workbook()
            self._write_measurements_sheet(workbook.active, session)
            self._write_info_sheet(workbook.create_sheet(_INFO_SHEET_NAME), session)
            workbook.save(str(path))
            return True
        except Exception:  # қорғаныс: болжанбаған қате де сыртқа шықпайды
            return False

    def _write_measurements_sheet(self, sheet: Worksheet, session: ExperimentSession) -> None:
        sheet.title = _MEASUREMENTS_SHEET_NAME

        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")

        for column_index, title in enumerate(_HEADER, start=1):
            cell = sheet.cell(row=1, column=column_index, value=title)
            cell.font = header_font
            cell.alignment = center_alignment

        start_time = session.measurements[0].timestamp
        for row_index, measurement in enumerate(session.measurements, start=2):
            elapsed = measurement.get_value("time")
            if elapsed is None:
                elapsed = (measurement.timestamp - start_time).total_seconds()

            row_number = row_index - 1
            number_cell = sheet.cell(row=row_index, column=1, value=row_number)
            number_cell.alignment = center_alignment

            time_cell = sheet.cell(row=row_index, column=2, value=elapsed)
            time_cell.number_format = _TIME_NUMBER_FORMAT
            time_cell.alignment = center_alignment

            for key, column in _VALUE_COLUMNS:
                value = measurement.get_value(key)
                cell = sheet.cell(row=row_index, column=column)
                if value is not None:
                    cell.value = value
                    cell.number_format = _VALUE_NUMBER_FORMAT
                cell.alignment = center_alignment

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for column_index, width in _COLUMN_WIDTHS.items():
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    def _write_info_sheet(self, sheet: Worksheet, session: ExperimentSession) -> None:
        rows: list[tuple[str, str]] = [
            ("Experiment ID", session.experiment_id),
            ("Start time", session.started_at.isoformat()),
        ]
        if session.ended_at is not None:
            rows.append(("End time", session.ended_at.isoformat()))
        rows.append(("Duration (s)", f"{session.duration_seconds():.2f}"))
        rows.append(("Measurement count", str(session.measurement_count)))

        label_font = Font(bold=True)
        for row_index, (label, value) in enumerate(rows, start=1):
            label_cell = sheet.cell(row=row_index, column=1, value=label)
            label_cell.font = label_font
            sheet.cell(row=row_index, column=2, value=value)

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 30
